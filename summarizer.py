#!/usr/bin/env python3
import os
import openai
import openpyxl 
from dotenv import load_dotenv
from datetime import datetime

class Resp(object):
    ''' An object for storing the OpenAI responses as attributes '''
    pass

def summ_prompt(ab_str):
    '''
    Given abstract paragraph string, 
    return the prompt string to request a summary.
    '''
    preamble = "Generate a one sentence summary of the following paragraph.  The summary should describe what research was conducted and the key conclusion."
    return preamble + "\n\nParagraph:\n" + ab_str

def app_prompt(ab_str):
    '''
    Given abstract string, 
    return the prompt string to request a list of potential applications.
    '''
    preamble = "Generate a list of two of the most relevant military and defense applications of the engineering technology described in the following paragraph.  Each item in the list in the list should be a single sentence.  The list should be formatted in markdown"
    return preamble + "\n\nParagraph:\n" + ab_str

def cat_prompt(categories, ab_str):
    '''
    Given a list of categories as strings and the abstract string,
    return the prompt to request categorization of the abstract.
    '''

    preamble = """Below is a set of categories and a one paragraph research abstract.  Select the one best category to fit the following paragraph.

"""
    for c in categories:
        preamble += c + '\n'
    
    return preamble + '\nParagraph:\n' +ab_str

def cat_prompt2(categories, title,  ab_str):
    '''
    Given a list of categories as strings and the abstract string,
    return the prompt to request categorization of the abstract.
    '''

    preamble = ('Here is a research project described by a title and abstract: \n\n Title: %s \n Abstract: %s \n And here is a list of categories: \n'%(title,ab_str))
    for c in categories:
        preamble += c + '\n'
    
    return preamble + 'Select the one category that best fits the research project.'


def ensure_newline(text):
    if not text.endswith('\n'):
        text += '\n'
    return text

# Load the .env variables and set API key.
load_dotenv() 
openai.api_key = os.getenv("OPENAI_API_KEY")

# Open Excel file with authors and abstracts.
datafile='abstracts2024.xlsx'
wb = openpyxl.load_workbook(datafile)
ws = wb.active

# Read rows,  skipping header, and generate list of Resp objects.
records = []
for row in ws.iter_rows(
        min_row = 2, max_row = ws.max_row,
        min_col=1, max_col=ws.max_column,
        values_only=True):
    resp = Resp()
    resp.name = row[2] + " " + row[1]
    resp.title = row[11]
    # Use thesis abstract as primary
    resp.abstract = row[12]
    if (resp.abstract is None):
        # try using proposal abstract
        resp.abstract = row[13]
    if (not (resp.abstract is None)):
        records.append(resp)
    else:
        print("! Can't find an abstract for %s"%resp.name)

# Specify the categories for organizing the topics.
categories = ['Materials Science and Engineering',
              'Robotics and Controls Engineering',
              'Space and Aerospace Engineering',
              'Thermodynamics, Fluid Dynamics and Energy',
              'Ship Design and Naval Engineering',
              'Survivability and Weaponeering',
              'Solid Mechanics and Structure Engineering']

# Specify which Open AI model we want to use.
model = "gpt-3.5-turbo-instruct"

# Loop through the records and do three queries for each record:
# 1. Summarize 2. Appply and 3. Categorize 
for ii in range(len(records)):
    print("Summarizing %d of %d for <%s>"%(ii+1,len(records),records[ii].name))
    psumm = summ_prompt(records[ii].abstract)
    rsumm = openai.Completion.create(model = model,
                                     prompt = psumm,
                                     max_tokens = 200,
                                     temperature=0.6)
    records[ii].summary = rsumm.choices[0].text

    papp = app_prompt(records[ii].abstract)
    rapp = openai.Completion.create(model = model,
                                     prompt = papp,
                                     max_tokens = 200,
                                     temperature=0.6)
    records[ii].applications = rapp.choices[0].text

    pcat = cat_prompt(categories, records[ii].abstract)
    rcat = openai.Completion.create(model = model,
                                     prompt = pcat,
                                     max_tokens = 200,
                                     temperature=0.6)
    records[ii].category = rcat.choices[0].text


# Using all the summaries, request a summary of summaries
ssumm = ''
for record in records:
    ssumm += record.summary


# Create a list of titles and summaries
sssumm = ''
for record, ii in zip(records,range(len(records))):
    sssumm += "*  Title: %s  \n Summary: %s \n\n"%(ii,record.title, record.summary)

# Generate some categories
prompt = ("Here is a list of research thesis titles and summaries: \n\n %s \n\n Provide a list of 5 naval-relevant research categories to summarize these research areas "% sssumm)

rcats = openai.Completion.create(model = model,
                                prompt = prompt,
                                max_tokens = 100,
                                temperature=0.3)
cats = rcats.choices[0].text.strip().split('\n')

# Re-categorize
for ii in range(len(records)):
    pcat = cat_prompt2(cats, records[ii].title, records[ii].abstract)
    rcat = openai.Completion.create(model = model,
                                     prompt = pcat,
                                     max_tokens = 200,
                                     temperature=0.6)
    records[ii].category = rcat.choices[0].text.strip()


prompt = 'Summarize the engineering topics described in the text below.  The summary should be a list of the five most important technology categories.  The list should be in markdown list format: \n\n' + ssumm 

rsumm= openai.Completion.create(model = model,
                                prompt = prompt,
                                max_tokens = 400,
                                temperature=0.3)

# Write the output as a markdown document
outf = 'synopsis.md'
print('Writing output to <%s>'%outf)
f = open(outf,'w')
now = datetime.now()
f.write("""# Synopsis of Abstracts

Generated summary of %d abstracts with the ``%s`` model on %s. \n"""%(len(records), model, now.strftime('%d %b %Y %H:%M:%S')))

f.write("\nEach abstract is summarized by a single sentence and two potential defense applications.\n")

f.write("The research can be summarized in the following categories:\n")
for c in cats:
    f.write(c)
    # Count the nubmer in this category
    cnt = 0
    for record in records:
        if record.category.find(c[3:]) >= 0:
            cnt += 1
    f.write(' [%d of %d]'%(cnt,len(records)))
    f.write('\n')

# Write the summaries for each category
for c in cats:
    f.write("\n\n## " + c + "\n")
    cnt = 0
    for record in records:
        if record.category.find(c[3:]) >= 0:
            f.write("\n\n### " + ensure_newline(record.name) + "\n")
            f.write("Title: %s"%ensure_newline(record.title))
            f.write(ensure_newline(record.summary))
            f.write("\n\n Naval Applications: %s"%ensure_newline(record.applications))
            cnt += 1
    if cnt < 1:
        f.write('\n--\n')

f.close()

